import openpyxl
import sys

wb = openpyxl.load_workbook('a_data.xlsx')
test = wb.get_sheet_by_name("Sheet1")
sent = ''

sys.stdout = open('output.txt','w', encoding="utf-8")

for i in range(3, 5000, 1) :
    sent += str(test.cell(row = i, column = 3).value)

print(sent)
#if test.cell(row = 2, cocumn = 1).value == '향토문화/음식' :





# print(test.cell(row = 1, column = 1).value)